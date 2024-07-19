import hashlib
import re

import openpyxl
from pymongo import MongoClient
from pymongo.errors import DuplicateKeyError

client = MongoClient('mongodb://readWrite:readWrite123456@180.76.162.140:27017/ks')
keys_coll = client['ks']['keys']

def hash_keyword(keyword):
    try:
        sum_str = 0
        for i in str(keyword):
            sum_str += ord(i)
        return int(sum_str%20)
    except:
        return 0

def check(_string):
    zh_model = re.compile(u'[\u4e00-\u9fa5]')
    match = zh_model.search(_string)
    if match:
        return True
    else:
        return False

def insert_keyword():
    # 打开Excel文件
    workbook = openpyxl.load_workbook(r'D:\Projects\kuaishou_android\整合关键词.xlsx')
    # workbook = openpyxl.load_workbook(r'D:\Projects\kuaishou_android\test.xlsx')

    # 选择要操作的工作表
    worksheet = workbook['Sheet1']

    # 获取工作表中的数据
    count = 0
    try:
        for row in worksheet.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    keyword = str(cell).strip()
                    number = hash_keyword(keyword)
                    if number == 0:
                        number = 20
                    hash_key = hashlib.md5(str(keyword).encode("utf8")).hexdigest()
                    if check(keyword):
                        _data = {
                            "hash_key": hash_key,
                            "keyword": keyword,
                            "number": number
                        }
                        try:
                            keys_coll.insert_one(_data)
                            print(f'正在入库第 {count} 个词， 对应手机编号为 {number} , 关键词是： {keyword}')
                        except DuplicateKeyError:
                            print(f"正在入库第 {count} 个词， 关键词：{keyword}  已经存在，无法入库")
                        except Exception as error:
                            print(error)
                    count +=1
    except Exception as error:
        print(error)


if __name__ == "__main__":
    insert_keyword()

